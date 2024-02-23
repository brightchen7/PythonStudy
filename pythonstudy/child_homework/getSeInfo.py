# 命名规则：数据类型_数据含义，字符串变量只有含义
# curr代表今天或者以今天为第一个日期，prev代表上一个交易日，next下一个
# 查询最新的、有收盘价的交易日，并定位为“今天”：函数QueryCurrTradingDay
# 如果某个计算行为需要X个交易日的数据，那么在这个计算函数一开始需要先过滤出有X个交易日的股票代码：函数code_list_filter
# 这个版本没有计算ExpMA的功能，待后续api上限再完善ExpMA
import concurrent.futures
import json
import os
import time
import datetime

import openpyxl
import pandas as pd
import requests
import urllib3
from openpyxl.styles import PatternFill

urllib3.disable_warnings()
token_valid_threshold = 3500

# 要用到的url和header统一放在这里
base_url = "http://10.110.13.221:8061"
query_trading_day_url = "/calendar/trading_day_query/ml"
get_ashare_url = "/data_service/get_ashare"
get_close_price_url = "/data_service/get_close_price"
get_period_close_price_url = "/data_service/get_period_close_price"
get_description_url = "/data_service/get_description"
queryIndustry_url = "/data_service/get_citics_industry"
headers = {
    "accept": "application/json",
    "Content-Type": "application/x-www-form-urlencoded",
}


def QueryCurrTradingDay():
    # 查询最新的有收盘价的交易日，作为curr_trade_date
    url = base_url + query_trading_day_url
    curr_day = datetime.datetime.now().strftime("%Y%m%d")
    body = {"base_day": curr_day, "offset": 0}
    response = requests.get(url, headers=headers, params=body)
    if curr_day == response.json()["message"]:
        # 如果今天就是交易日，那么判断当前时刻是否早于15：50
        if datetime.datetime.now().time() < datetime.time(hour=15, minute=50, second=0):
            body["offset"] = -1
            res = requests.get(url=url, headers=headers, params=body)
            return res.json()["message"]
        else:
            return response.json()["message"]
    else:
        # 如果今天不是交易日，那么offset传0即可
        body["offset"] = 0
        res = requests.get(url=url, headers=headers, params=body)
        return res.json()["message"]


def QueryTradingDay(exact_day, offset):
    # exact_day是标准日，offset是偏置，求偏置的交易日，如果输入是list，则返回多个offset对应的交易日
    url = base_url + query_trading_day_url
    body = {"base_day": exact_day, "offset": 0}
    if isinstance(offset, int):
        body["offset"] = offset
        response = requests.get(url=url, headers=headers, params=body)
        return response.json()["message"]
    elif isinstance(offset, list):
        list_trading_day = []
        for day_offset in offset:
            body["offset"] = day_offset
            response = requests.get(url=url, headers=headers, params=body)
            list_trading_day.append(response.json()["message"])
        return list_trading_day
    else:
        print("offset must be list<int> or int")
        return None


# 当前交易日被认为是具有收盘价的最新交易日日期
curr_trade_date = QueryCurrTradingDay()
# 以curr_trade_date为基准的上个交易日
prev_trade_date = QueryTradingDay(curr_trade_date, -1)
# 以curr_trade_date为基准的下个交易日
next_trade_date = QueryTradingDay(curr_trade_date, 1)

# 以curr_trade_date为基准的往前一共十个交易日 获取每天的
list_currTen_offset = [i for i in range(-9, 1)]  # [-9,-8,...,-1,0]
list_currTen = QueryTradingDay(curr_trade_date, list_currTen_offset)

# 以prev_trade_date为基准的的往前一共十个交易日 获取每天的
list_prevTen_offset = [i for i in range(-10, 0)]  # [-10,-8,...,-2,-1]
list_prevTen = QueryTradingDay(curr_trade_date, list_prevTen_offset)

# # 以curr_trade_date为基准的往前一共二十个交易日 获取每天的 用于计算ExpMA20_daily
# list_currTwty_offset = [i for i in range(-19, 1)]  # [-19,-18,...,-1,0]
# list_currTwty = QueryTradingDay(curr_trade_date, list_currTwty_offset)
#
# # 以curr_trade_date为基准的往前一共一百个交易日的每五日末尾 用于计算ExpMA20_weekly
# list_currHdrd_offset = [i for i in range(-95, 1, 5)]  # [-95,-90,...,-5,0]
# list_currHdrd = QueryTradingDay(curr_trade_date, list_currHdrd_offset)


def code_list_filter(list_tickers_code, days_num):
    # 这是根据days_num过滤股票代码的函数，如果某个计算行为需要过去30日交易日的股票，则需要靠这个函数筛选出具有连续30个交易日的股票代码
    # 不能根据上市日过滤股票代码，因为期间可能发生停牌和复牌，按照开始日和结束日的收盘价数据数量够不够来过滤
    # QueryTradingDay传入的offset = -1*(days_num - 1)，得到的开始日期到curr_trade_date期间一共days_num个交易日
    # 拿到要求的最早交易日日期，如果earliest ~ curr之间股票收盘价的数量不足days_num，则过滤掉
    earliest_trading_date = QueryTradingDay(curr_trade_date, -1 * (days_num - 1))
    patch_amount = int(10000/days_num)  # 1<patch_amount<2000，兼顾过滤函数的运行速度和内存处理能力等因素
    url = base_url + get_period_close_price_url

    def sub_code_list_filter(list_sub):
        set_filtered = set()
        body = {
            "tickers": list_sub,
            "start_date": earliest_trading_date,
            "end_date": curr_trade_date,
        }
        response = requests.post(url, headers=headers, data=body)
        for key in response.json():
            if len(response.json()[key]) == days_num:
                set_filtered.add(key)
        return set_filtered

    list_subCode = [
        list_tickers_code[i: i + patch_amount]
        for i in range(0, len(list_tickers_code), patch_amount)
    ]
    with concurrent.futures.ThreadPoolExecutor() as executor:
        results = list(executor.map(sub_code_list_filter, list_subCode))
    # results = [{set_filtered},{},{},{},]
    set_final_filtered = set()
    for set_result in results:
        set_final_filtered.update(set_result)
    return list(set_final_filtered)


def Wind2Code(wind):
    if "." in wind:
        number, market = wind.split(".")
        code = market + number
        return code
    else:
        return wind


def flt2pct(num):
    # 将浮点数转换为百分之多少多少
    percentage = num * 100
    formatted_percentage = "{:.2f}%".format(percentage)
    return formatted_percentage


def get_ashare():
    # 首先获得所有A股的股票代码和简称 get
    url = base_url + get_ashare_url
    params = {"trade_date": f"{curr_trade_date}"}
    response = requests.get(url=url, params=params, headers=headers)
    return response.json()


def get_close_price(tickers, targDate):
    # 获取当天收盘价 post
    url = base_url + get_close_price_url
    body = {"tickers": tickers, "trade_date": f"{targDate}"}
    response = requests.post(url, headers=headers, data=body)
    return response.json()  # [{},]


def MA_daily(list_tickers_code, list_days):
    # 传入股票代码和要查询的交易日的list
    list_tickers_code = code_list_filter(list_tickers_code, len(list_days))
    # 获取多日平均收盘价 post
    url = base_url + get_period_close_price_url
    body = {
        "tickers": list_tickers_code,
        "start_date": f"{list_days[0]}",
        "end_date": f"{list_days[-1]}",
    }
    response = requests.post(url, headers=headers, data=body)
    # response.json()是字典，key是wind代码，value是连续多日收盘价的list，求平均后输出为字典
    return {
        key: round(sum(values) / len(values), 3)
        for key, values in response.json().items()
    }


def QueryIndustry(list_tickers_code):
    url = base_url + queryIndustry_url
    body = {"tickers": list_tickers_code}
    response = requests.post(url, headers=headers, data=body)
    return {
        item["S_INFO_WINDCODE"]: item["CITICS_IND_NAME"]
        for item in response.json()
        if item is not None
    }


def QuerySector(code):
    # 这个规律是根据股票代码在东方财富上的所在板块总结的
    if code.startswith("6"):
        if code.startswith("60"):
            return "上交所主板"
        elif code.startswith("688"):
            return "上交所科创板"
        else:
            return "上交所未知板块"
    elif code.startswith("0"):
        if code.startswith("000"):
            return "深交所主板"
        if code.startswith("001"):
            return "深交所主板"
        elif code.startswith("002"):
            return "深交所中小板"
        elif code.startswith("003"):
            return "深交所中小板"
        else:
            return "深交所未知板块"
    elif code.startswith("30"):
        return "深交所创业板"
    elif code.startswith("8"):
        return "北交所"
    elif code.startswith("4"):
        return "北交所"
    else:
        return "未知板块"


def WaveQuery(list_tickers_code, days_num):
    # 查询days_num之前的那天交易日的close，然后在表格内进行字段间运算
    list_days_numCode = code_list_filter(list_tickers_code, days_num)
    targ_trade_date = QueryTradingDay(curr_trade_date, -1 * (days_num - 1))
    list_targDay_close = get_close_price(list_days_numCode, targ_trade_date)
    list_currDay_close = get_close_price(list_days_numCode, curr_trade_date)
    return {
        item["S_INFO_WINDCODE"]:
            flt2pct(
                round(
                    list_currDay_close[index]["CLOSE_PRICE"] / list_targDay_close[index]["CLOSE_PRICE"] - 1,
                    4,
                )
            )
        for index, item in enumerate(list_targDay_close)
    }


def colColor(filePath, colNameList, colorStr):
    # 传入excel文件路径，列名的list，要转换的颜色RGB十六进制字符串
    wb = openpyxl.load_workbook(filePath)
    sheet = wb.active
    for item in colNameList:
        column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == item:
                    column_index = cell.column_letter
                    break
        if column_index is None:
            break
        fill = PatternFill(start_color=colorStr, end_color=colorStr, fill_type="solid")
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=sheet[column_index][0].column,
            max_col=sheet[column_index][-1].column,
        ):
            for cell in row:
                cell.fill = fill
    wb.save(filePath)


class CoremailHandler:
    def __init__(self, url, uid, password):
        self.sid = None
        self.session = None
        self.cookies = None
        self.base_url = url
        self.login_success = False
        self.login_time = None
        self._uid = uid
        self._pass = password

    def login(self):
        now_time_int = int(time.time())
        if (
            not self.login_time
            or now_time_int > self.login_time + token_valid_threshold
        ):
            login_url = f"{self.base_url}user:login"
            data = json.dumps({"uid": self._uid, "password": self._pass})
            header = {"Content-Type": "text/x-json"}
            response = requests.post(
                login_url, headers=header, data=data, verify=False
            )
            self.cookies = response.cookies
            text = json.loads(response.text)
            if text.get("code") == "S_OK":
                self.sid = text["var"]["sid"]
                self.login_success = True
                self.login_time = now_time_int
            else:
                print("Attention:Fail to login.")

    def _req(self, url, data, **kwargs):
        self.login()
        params = {
            "headers": {"Content-Type": "text/x-json"},
            "cookies": self.cookies,
            "verify": False,
        }
        if data is not None:
            params["data"] = json.dumps(data)
        for key, value in kwargs.items():
            params[key] = value
        response = requests.post(f"{self.base_url}{url}&sid={self.sid}", **params)
        text = json.loads(response.text)
        if text.get("code") == "S_OK":
            return text["var"]
        else:
            print(f"Attention:Fail to request result:-->{text}, response:-->{response}")
            return []

    def _get(self, url, data, **kwargs):
        self.login()
        params = {
            "cookies": self.cookies,
            "verify": False,
        }
        for key, value in kwargs.items():
            params[key] = value
        response = requests.get(
            f"{self.base_url}{url}&sid={self.sid}", params=data, **params
        )
        text = json.loads(response.text)
        if text.get("code") == "S_OK":
            return text["var"]
        else:
            print(f"Attention:Fail to request result:-->{text}, response:-->{response}")
            return None

    def _req_download(self, data, filename):
        self.login()
        response = requests.get(
            f"{self.base_url}mbox:getMessageData&sid={self.sid}",
            data,
            cookies=self.cookies,
            stream=True,
            verify=False,
        )
        if response.status_code == 200:
            if filename:
                with open(filename, "wb") as f:
                    f.write(response.content)
                print(f"Save File to:-->{filename}")
                return True
            else:
                return response.text
        else:
            print(
                f"Attention:Fail to request result:-->{response.text}, response:-->{response}"
            )
            return False

    def getAllFolders(self):
        url = "mbox:getAllFolders"
        res = self._req(
            url,
            {"flush": True, "stats": True, "threads": True},
        )
        return res

    @staticmethod
    def _deep_search_fold(item, name):
        if item["name"] == name:
            return item["id"], item.get("children")
        if item and item.get("children"):
            children_folder = item["children"]
            for folder in children_folder:
                if folder["name"] == name:
                    return folder["id"], folder.get("children")
        return None, None

    def _recFolder(self, item, name):
        names = name.split(".")
        for i in range(len(names)):
            folder_id, folder_children = self._deep_search_fold(item, names[i])
            item = folder_children
        return folder_id

    def getFolderIdByName(self, name):
        if name == "收件箱":
            return 1
        else:
            all_folders = self.getAllFolders()
            folder_id = None
            for item in all_folders:
                if item["name"] == name:
                    folder_id = self._recFolder(item, name)
            return folder_id

    def listMessages(
        self, fid, received_date=None, start=0, limit=-1, summaryWindowSize=0
    ):
        url = "mbox:listMessages"
        params = {
            "fid": fid,
            "start": start,
            "limit": limit,
            "summaryWindowSize": summaryWindowSize,
            "returnTotal": "true",
        }
        if received_date:
            params.update({"filter": {"receivedDate": received_date}})
        res = self._req(url, params)
        return res

    def readMessage(self, ID):
        return self._req("mbox:readMessage", {"id": ID, "mode": "text"})

    def getMessageData(self, ID, partId=0, filename=None):
        return self._req_download({"mid": ID, "part": str(partId)}, filename)

    def getMessageAttachmentsAll(self, ID, folderPath):
        message = self.readMessage(ID)
        final_res = []
        for item in message["attachments"]:
            res = self.getMessageData(
                ID, item["id"], f'{folderPath}/{item["filename"]}'
            )
            final_res.append(res)
        return final_res

    def getMessagePartialAttachment(self, ID, partId, fileName):
        return self.getMessageData(ID, partId, fileName)

    def getHTML(self, ID):
        message = self.readMessage(ID)
        item = message["html"]
        html = self.getMessageData(ID, item["id"])
        return html

    def prepareCompose(self):
        return self._get("mbox:compose", {})

    def prepareAttachment(self, composeId, attachment):
        return self._req(
            url="upload:prepare",
            data={
                "composeId": composeId,
                "fileName": os.path.basename(attachment),
                "size": 10,
            },
        )

    def uploadAttachment(self, composeId, attachmentRes, attachment):
        file_name = os.path.basename(attachment)
        attachment_id = attachmentRes["attachmentId"]
        url = f"upload:directData&composeId={composeId}&attachmentId={attachment_id}"
        return self._req(
            url,
            data=None,
            files={"file": (file_name, open(attachment, "rb"))},
            headers={},
        )

    def sendEmail(self, data):
        result = self._req(url="mbox:compose", data=data)
        print("Email Sent Success!")
        return result

    def searchEmailList(
        self, folder="收件箱", start=0, limit=10, summaryWindowSize=10, pattern=""
    ):
        fid = self.getFolderIdByName(name=folder)
        if pattern:
            data = {
                "fid": fid,
                "start": start,
                "limit": limit,
                "pattern": pattern,
                "summaryWindowSize": summaryWindowSize,
                "order": "date",
                "desc": True,
                "conditions": [],
                "groupings": {},
            }
            return self._req("mbox:searchMessages", data)
        else:
            return self.listMessages(
                fid, start=start, limit=limit, summaryWindowSize=summaryWindowSize
            )


class CoremailHelper:
    def __init__(self, uid, password, testing=False):
        url = "https://mailbj.cicc.group/coremail/s/json?func="
        self.domain = "cicc.com.cn"
        if testing:
            url = "https://bjmail.cicccs.group/coremail/s/json?func="
            self.domain = "cs.cicc.com.cn"
        self.__handler = CoremailHandler(url, uid, password)

    def get_email_addresses(self, addrStr=""):
        if addrStr:
            addr_list = addrStr.split(";")
            for i in range(0, len(addr_list)):
                if "@" not in addr_list[i]:
                    addr_list[i] = addr_list[i] + "@" + self.domain
            return addr_list
        return []

    def send_mail(
        self,
        recipients="",
        cc="",
        body="",
        behalf="",
        subject="",
        attachments=None,
        auto=False,
        attrs=None,
        bcc="",
    ):
        compose_id = self.__handler.prepareCompose()
        data = {
            "id": compose_id,
            "attrs": {
                "to": recipients.split(";"),
                "cc": cc.split(";"),
                "bcc": bcc.split(";"),
                "subject": subject,
                "content": body,
                "isHtml": True,
            },
            "returnInfo": True,
        }
        if attrs:
            for key, value in attrs.items():
                data["attrs"][key] = value
        if behalf:
            data["attrs"]["account"] = behalf
            data["attrs"]["from"] = behalf
        if attachments is not None and len(attachments) > 0:
            attachments_payload = []
            for attachment in attachments:
                attachment_res = self.__handler.prepareAttachment(
                    compose_id, attachment
                )
                upload_res = self.__handler.uploadAttachment(
                    compose_id, attachment_res, attachment
                )
                attachments_payload.append(
                    {
                        "id": upload_res["attachmentId"],
                        "type": "upload",
                        "name": upload_res["fileName"],
                    }
                )
            data["attachments"] = attachments_payload
        if auto is True:
            data["action"] = "deliver"
        else:
            data["action"] = "save"

        return self.__handler.sendEmail(data)

    def search_email(
        self, folder="收件箱", start=0, limit=10, summaryWindowSize=10, pattern=""
    ):
        return self.__handler.searchEmailList(
            folder=folder,
            start=start,
            limit=limit,
            summaryWindowSize=summaryWindowSize,
            pattern=pattern,
        )

    def get_email_info(self, ID):
        mail_id = ID
        if isinstance(ID, dict):
            mail_id = ID["id"]
        return self.__handler.readMessage(mail_id)

    def download_email(self, ID, folder):
        mail_id = ID
        if isinstance(ID, dict):
            mail_id = ID["id"]
        email_info = self.__handler.readMessage(mail_id)
        return self.__handler.getMessageData(
            ID, 0, f'{folder}\\{email_info["subject"]}.eml'
        )

    def download_email_all_attachments(self, ID, folder):
        mail_id = ID
        if isinstance(ID, dict):
            mail_id = ID["id"]
        return self.__handler.getMessageAttachmentsAll(mail_id, folder)

    def download_email_content(self, ID, folder):
        mail_id = ID
        if isinstance(ID, dict):
            mail_id = ID["id"]
        email_info = self.__handler.readMessage(mail_id)
        item = email_info["html"]
        return self.__handler.getMessagePartialAttachment(
            mail_id, item["id"], f'{folder}/{email_info["subject"]}.html'
        )

    def download_email_attachment(self, ID, attachmentId, folder):
        mail_id = ID
        if isinstance(ID, dict):
            mail_id = ID["id"]
        email_info = self.__handler.readMessage(mail_id)
        if email_info.get("attachments") and len(email_info["attachments"]) > 0:
            for attachment in email_info["attachments"]:
                if attachment["id"] == attachmentId:
                    return self.__handler.getMessagePartialAttachment(
                        mail_id, attachment["id"], f'{folder}/{attachment["filename"]}'
                    )


if __name__ == "__main__":
    final_file_name = f"./股票汇总数据_{curr_trade_date}.xlsx"

    # 获得A股代码和简称的字典，然后转为dataframe
    # df_final是最后要输出为excel的dataframe，后面所有拿到的数据都要与df_final.left_merge on "万得代码"
    dict_code_name = get_ashare()
    df_final = pd.DataFrame(list(dict_code_name.items()), columns=["万得代码", "股票简称"])
    list_code = list(dict_code_name.keys())

    print("获取当前收盘价...")
    # 当前收盘价
    list_close_price = get_close_price(list_code, curr_trade_date)
    df_close_price = pd.DataFrame(list_close_price)
    df_final = pd.merge(
        df_final,
        df_close_price[["S_INFO_WINDCODE", "CLOSE_PRICE"]],
        left_on="万得代码",
        right_on="S_INFO_WINDCODE",
        how="left",
    )
    df_final = df_final.drop(columns="S_INFO_WINDCODE")
    df_final = df_final.rename(columns={"CLOSE_PRICE": "现价(元)"})

    print("获取十日滑动平均...")
    # 昨日的十日MA
    # dict_preTen_MA的key是股票代码，val是昨日的十日MA
    dict_preTen_MA = MA_daily(list_code, list_prevTen)
    df_preTen_MA = pd.DataFrame(
        list(dict_preTen_MA.items()), columns=["万得代码", f"{prev_trade_date}"]
    )
    df_final = pd.merge(df_final, df_preTen_MA, on="万得代码", how="left")
    # 今日MA
    dict_currTen_MA = MA_daily(list_code, list_currTen)
    df_currTen_MA = pd.DataFrame(
        list(dict_currTen_MA.items()), columns=["万得代码", f"{curr_trade_date}(今天MA10)"]
    )
    df_final = pd.merge(df_final, df_currTen_MA, on="万得代码", how="left")
    # 明日MA=3*今日MA-2*昨日MA
    df_final[f"{next_trade_date}"] = df_final.apply(
        lambda row: round(
            (row[f"{curr_trade_date}(今天MA10)"] * 3 - row[f"{prev_trade_date}"] * 2), 3
        ),
        axis=1,
    )

    # 距离目标价
    df_final["MA10距离目标价"] = df_final.apply(
        lambda row: flt2pct(round((1 - row[f"{next_trade_date}"] / row["现价(元)"]), 4))
        if pd.notnull(row[f"{next_trade_date}"]) and pd.notnull(row["现价(元)"])
        else None,
        axis=1,
    )

    # 所属行业
    dict_industry = QueryIndustry(list_code)
    df_industry = pd.DataFrame(list(dict_industry.items()), columns=["万得代码", "所属行业"])
    df_final = pd.merge(df_final, df_industry, on="万得代码", how="left")

    # 601995.SH → SH601995
    print("获取Code...")
    df_final["Code"] = df_final.apply(lambda row: Wind2Code(row["万得代码"]), axis=1)

    # 所属板块
    print("获取Sector...")
    df_final["Sector"] = df_final.apply(lambda row: QuerySector(row["万得代码"]), axis=1)

    # 最近5/10/20/30日涨跌幅
    print("获取5/10/20/30日涨跌幅...")
    days = 5
    dict_currWave = WaveQuery(list_code, days)
    df_currWave = pd.DataFrame(
        list(dict_currWave.items()), columns=["万得代码", f"{days}日涨跌幅"]
    )
    df_final = pd.merge(df_final, df_currWave, on="万得代码", how="left")

    days = 10
    dict_currWave = WaveQuery(list_code, days)
    df_currWave = pd.DataFrame(
        list(dict_currWave.items()), columns=["万得代码", f"{days}日涨跌幅"]
    )
    df_final = pd.merge(df_final, df_currWave, on="万得代码", how="left")

    days = 20
    dict_currWave = WaveQuery(list_code, days)
    df_currWave = pd.DataFrame(
        list(dict_currWave.items()), columns=["万得代码", f"{days}日涨跌幅"]
    )
    df_final = pd.merge(df_final, df_currWave, on="万得代码", how="left")

    days = 30
    dict_currWave = WaveQuery(list_code, days)
    df_currWave = pd.DataFrame(
        list(dict_currWave.items()), columns=["万得代码", f"{days}日涨跌幅"]
    )
    df_final = pd.merge(df_final, df_currWave, on="万得代码", how="left")

    df_final.to_excel(final_file_name, index=False)
    print(f"{final_file_name} 输出成功")

    # # 可选功能：着色，直接将要变色的列名写入toXXX的list，指定RGB的六位十六进制码的字符串
    # # 将列名的list和RGB字符串传入colColor
    # print("着色...")
    # # RGB转换为FF0000
    # toFF0000 = ['万得代码', '股票简称', '现价(元)']
    # colColor(final_file_name, toFF0000, 'FF0000')

    # 以下内容是coremail的邮件收发
    # 发送邮件，将上文的Excel以附件发送给指定邮箱
    print("邮件发送中...")
    senderParams = {
        "uid": "chenl@cicc.com.cn",
        "password": "1q2w#E$R%T^Y",
        "testing": False,
    }
    email_receiver = "bright_chen7@163.com"
    attachment_files = [final_file_name, ]
    email_subject = "股票汇总数据"
    sender = CoremailHelper(**senderParams)
    sender.send_mail(
        recipients=email_receiver,
        cc="",
        body="股票汇总数据",
        subject=email_subject,
        attachments=attachment_files,
        auto=True,
    )
    print("邮件发送成功")

    # # 接收者下载附件：创建download_to的文件夹并下载上文Excel附件
    # receiverParams = {
    #     "uid": "yug.wang@cs.cicc.com.cn",
    #     "password": "xvT!672qct",
    #     "testing": True,
    # }
    # receiver = CoremailHelper(**receiverParams)
    # download_to = "./download"
    # # 自定义下载目录，直接修改download_to
    # if not os.path.exists(download_to):
    #     os.mkdir(download_to)
    # email_subject = "股票汇总数据"
    # inbox_email_id_list = receiver.search_email(pattern=email_subject)
    # if inbox_email_id_list and len(inbox_email_id_list):
    #     latest_email_id = inbox_email_id_list[0]
    #     receiver.download_email_all_attachments(latest_email_id, download_to)
    # else:
    #     print(f"找不到匹配对象：{email_subject}")
    # print(f"附件下载到：{download_to}")
